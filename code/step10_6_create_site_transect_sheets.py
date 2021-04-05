#!/usr/bin/env python
"""
Copyright 2021 Robert McGregor

Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation the
rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software,
and to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions of the
Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE
WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NON INFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR
COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""

import xlsxwriter
from openpyxl import load_workbook
import pandas as pd


def create_worksheet_fn(workbook, worksheet_name):
    """ Create establishment worksheet and set row height.
            :param workbook: open workbook object derived from create_workbook_fn function.
            :param worksheet_name: string object containing worksheet name."""

    # Set up Site Establishment worksheet #
    worksheet = workbook.add_worksheet(worksheet_name)

    # Set row height
    worksheet.set_default_row(56.25)

    return workbook, worksheet


def insert_sheet_headings_fn(workbook, worksheet, heading2, heading4, color_fill, range_value):
    """ Add item headings to cells as strings.
            :param workbook: open workbook object derived from create_workbook_fn function.
            :param worksheet: worksheet object current worksheet derived from create_worksheet_fn.
            :param heading2: workbook style derived  from define heading2_fn.
            :param heading4: workbook style derived  from define heading4_fn.
            :param color_fill:
            :param range_value:
            :return: workbook: updated workbook object.
            :return worksheet: updated worksheet object."""

    worksheet.write_string('B2', 'TRANSECT 1', heading2)
    worksheet.write_string('C2', 'START POINT', heading2)
    worksheet.write_string('B3', 'GROUND LAYER', heading4)
    worksheet.write_string('C3', 'BELOW', heading4)
    worksheet.write_string('D3', 'ABOVE', heading4)
    worksheet.write_column('A4', range_value, heading4)
    worksheet.write('A1', None, color_fill)
    worksheet.write('A2', None, color_fill)
    worksheet.write('A3', None, color_fill)

    return workbook, worksheet, range_value


def insert_blank_formatted_cells_fn(workbook, worksheet, heading7, range_value):
    """ Add blank formatted cells to worksheet.
            :param range_value:
            :param workbook: open workbook object derived from create_workbook_fn function.
            :param worksheet: worksheet object current worksheet derived from create_worksheet_fn.
            :param heading7: workbook style derived  from define heading7_fn.
            :return: workbook: updated workbook object.
            :return worksheet: updated worksheet object."""

    worksheet.write_column('B4', range_value, heading7)
    worksheet.write_column('C4', range_value, heading7)
    worksheet.write_column('D4', range_value, heading7)

    return workbook, worksheet


def merge_cells(workbook, worksheet, heading1, heading7):
    """ Add item headings to cells and merge.
            :param workbook: open workbook object derived from create_workbook_fn function.
            :param worksheet: worksheet object current worksheet derived from create_worksheet_fn.
            :param heading1: workbook style derived  from define heading1_fn.
            :param heading7: workbook style derived  from define heading7_fn.
            :return: workbook: updated workbook object.
            :return worksheet: updated worksheet object."""

    worksheet.merge_range('B1:E1', 'STEP 4A - TRANSECT 1', heading1)
    worksheet.merge_range('D2:E2', '', heading7)

    return workbook, worksheet


def define_column_widths_fn(workbook, worksheet):
    """ define and set column widths.
            :param workbook: open workbook object derived from create_workbook_fn function.
            :param worksheet: worksheet object current worksheet derived from create_worksheet_fn.
            :return: workbook: updated workbook object.
            :return worksheet: updated worksheet object."""

    worksheet.set_column('A:A', 9.09)
    worksheet.set_column('B:B', 48.00)
    worksheet.set_column('C:C', 48.00)
    worksheet.set_column('D:D', 48.00)
    worksheet.set_column('E:E', 13.00)

    return workbook, worksheet


def extract_features_to_list_fn(clean_df_list):
    """ Extract ground, below and above variables and add them to a list of lists.
            :param clean_df_list:
            :return content_list:  """

    df = clean_df_list[0]

    content1 = df.ground.tolist()
    content2 = df.below.tolist()
    content3 = df.above.tolist()

    content_list = [content1, content2, content3]

    return content_list


def insert_variables_to_transect_sheet_fn(content_list, workbook, worksheet):
    """ Insert df values to Observational Sheet (column by column).
            :param content_list:
            :param workbook:
            :param worksheet:
            :return: workbook: updated workbook object.
            :return worksheet: updated worksheet object."""

    # define starting column value
    col = 0

    for content in content_list:
        # define row and column start positions (loop)
        row = 3
        col += 1

        for item in content:
            # write list to column
            worksheet.write(row, col, item)
            # increase row value by 1
            row += 1

    return workbook, worksheet


def main_routine(clean_df_list, color_fill, heading1, heading2, heading4, heading7, workbook, star,
                 insert_vertical_data_fn):
    """
            :param clean_df_list:
            :param color_fill:
            :param heading1:
            :param heading2:
            :param heading4:
            :param heading7:
            :param workbook:
            :param star:
            :param insert_vertical_data_fn:
    """
    worksheet_name_list = ['Step 4A - Transect 1', 'Step 4B - Transect 2', 'Step 4C - Transect 3']

    x = 0
    for i in worksheet_name_list:

        worksheet_name = i

        range_value = range(1, 101)

        # call the create_worksheet_fn function.
        workbook, worksheet = create_worksheet_fn(workbook, worksheet_name)

        # call the insert_sheet_headings_fn function.
        workbook, worksheet, range_value = insert_sheet_headings_fn(workbook, worksheet, heading2, heading4, color_fill,
                                                                    range_value)

        # call the insert_blank_formatted_cells_fn function.
        workbook, worksheet = insert_blank_formatted_cells_fn(workbook, worksheet, heading7, range_value)

        # call the merge_cells_fn function.
        workbook, worksheet = merge_cells(workbook, worksheet, heading1, heading7)

        # call the define_column_widths_fn(workbook) function.
        workbook, worksheet = define_column_widths_fn(workbook, worksheet)

        # call the extract_features_to_list_fn function.
        content_list = extract_features_to_list_fn(clean_df_list)

        # call the insertVariablesToTransect1Sheet function.
        insert_variables_to_transect_sheet_fn(content_list, workbook, worksheet)
        x += 1
        # call the insertDataFN function
        if [star['transect' + str(x)][0]]:
            insert_vertical_data_fn(worksheet, 1, 3, [star['transect' + str(x)][0]], heading7, 1)

        print('Transect: ', x, ' - step10_6_create_site_transect sheets.py COMPLETE.')
    print("step10_7_create_site_basal_sheet.py initiating..........")


if __name__ == '__main__':
    main_routine()
