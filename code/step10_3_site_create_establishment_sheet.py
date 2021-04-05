#!/usr/bin/env python

"""
Copyright 2021 Robert McGregor

Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation the
rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software,
and to permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions of the
Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE
WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NON INFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR
COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.

"""

# import modules
import xlsxwriter
from openpyxl import load_workbook
import pandas as pd


def create_worksheet_fn(workbook, worksheet_name):
    """ Create establishment worksheet and set row height.
            :param workbook: open workbook object derived from create_workbook_fn function.
            :param worksheet_name: string object containing worksheet name."""

    # Set up Site Establishment worksheet
    worksheet = workbook.add_worksheet(worksheet_name)

    # Set row height
    worksheet.set_default_row(56.25)

    return workbook, worksheet


def insert_sheet_headings_fn(workbook, worksheet, heading2, heading3):
    """ Add item headings to cells as strings.
            :param workbook: open workbook object derived from create_workbook_fn function.
            :param worksheet: worksheet object current worksheet derived from create_worksheet_fn.
            :param heading2: workbook style derived  from define heading2_fn.
            :param heading3: workbook style derived  from define heading3_fn.
            :return: workbook: updated workbook object.
            :return worksheet: updated worksheet object."""

    worksheet.write_string('A2', 'ITEM', heading2)
    worksheet.write_string('A3', 'RECORDER:', heading3)
    worksheet.write_string('A4', 'ESTIMATOR:', heading3)
    worksheet.write_string('A5', 'ANY OTHERS PRESENT:', heading3)
    worksheet.write_string('A6', 'PROPERTY:', heading3)
    worksheet.write_string('A7', 'UNLISTED PROPERTY: (if required)', heading3)
    worksheet.write_string('A8', 'SITE ID:', heading3)
    worksheet.write_string('A9', 'PADDOCK NAME:', heading3)
    worksheet.write_string('A10', 'DATE & TIME:', heading3)
    worksheet.write_string('A11', 'DIRECTION FROM OFFSET:', heading3)
    worksheet.write_string('A12', 'DATUM:', heading3)
    worksheet.write_string('A13', 'OFFSET (LAT):', heading3)
    worksheet.write_string('A14', 'OFFSET (LONG):', heading3)
    worksheet.write_string('A15', 'CENTRE POINT (LAT):', heading3)
    worksheet.write_string('A16', 'CENTRE POINT (LONG):', heading3)
    worksheet.write_string('A17', 'LANDSCAPE POSITION:', heading3)
    worksheet.write_string('A18', 'SOIL SURFACE COLOUR:', heading3)
    worksheet.write_string('A19', 'SITE DESCRIPTION:', heading3)
    worksheet.write_string('A20', 'REASON FOR SITE SELECTION:', heading3)
    worksheet.write_string('A21', 'LAND SYSTEM:', heading3)
    worksheet.write_string('D21', 'SOURCE', heading3)
    worksheet.write_string('A22', 'CONSISTENT WITH MAPPING?:', heading3)
    worksheet.write_string('D22', 'IF NO, WHAT LAND SYSTEM:', heading3)
    worksheet.write_string('A23', 'NOTES:', heading3)
    worksheet.write_string('A24', 'NEAREST STOCK WATER (NAME):', heading3)
    worksheet.write_string('D24', 'NEAREST STOCK WATER (TYPE):', heading3)
    worksheet.write_string('A25', 'DISTANCE FROM NEAREST STOCK WATER (km):', heading3)
    worksheet.write_string('D25', 'DIRECTION FROM NEAREST STOCK WATER:', heading3)
    worksheet.write_string('A26', 'DISTANCE FROM TRACK (m):', heading3)
    worksheet.write_string('D26', 'DIRECTION FROM TRACK:', heading3)
    worksheet.write_string('A27', 'OTHER NEARBY INFRASTRUCTURE:', heading3)

    return workbook, worksheet


def insert_blank_formatted_cells_fn(workbook, worksheet, heading7):
    """ Add blank formatted cells to worksheet.
            :param workbook: open workbook object derived from create_workbook_fn function.
            :param worksheet: worksheet object current worksheet derived from create_worksheet_fn.
            :param heading7: workbook style derived  from define heading7_fn.
            :return: workbook: updated workbook object.
            :return worksheet: updated worksheet object."""

    worksheet.write_blank('E21', None, heading7)
    worksheet.write_blank('E22', None, heading7)
    worksheet.write_blank('E24', None, heading7)
    worksheet.write_blank('E25', None, heading7)
    worksheet.write_blank('E26', None, heading7)

    return workbook, worksheet


def merge_cells(workbook, worksheet, worksheet_name, heading1, heading2, heading7):
    """ Add item headings to cells and merge.
            :param workbook: open workbook object derived from create_workbook_fn function.
            :param worksheet: worksheet object current worksheet derived from create_worksheet_fn.
            :param worksheet_name: string object containing the worksheet name.
            :param heading1: workbook style derived  from define heading1_fn.
            :param heading2: workbook style derived  from define heading2_fn.
            :param heading7: workbook style derived  from define heading7_fn.
            :return: workbook: updated workbook object.
            :return worksheet: updated worksheet object."""

    worksheet.merge_range('A1:E1', worksheet_name, heading1)
    worksheet.merge_range('B2:E2', 'INPUT', heading2)
    worksheet.merge_range('B3:E3', '', heading7)
    worksheet.merge_range('B4:E4', '', heading7)
    worksheet.merge_range('B5:E5', '', heading7)
    worksheet.merge_range('B6:E6', '', heading7)
    worksheet.merge_range('B7:E7', '', heading7)
    worksheet.merge_range('B8:E8', '', heading7)
    worksheet.merge_range('B9:E9', '', heading7)
    worksheet.merge_range('B10:E10', '', heading7)
    worksheet.merge_range('B11:E11', '', heading7)
    worksheet.merge_range('B12:E12', '', heading7)
    worksheet.merge_range('B13:E13', '', heading7)
    worksheet.merge_range('B14:E14', '', heading7)
    worksheet.merge_range('B15:E15', '', heading7)
    worksheet.merge_range('B16:E16', '', heading7)
    worksheet.merge_range('B17:E17', '', heading7)
    worksheet.merge_range('B18:E18', '', heading7)
    worksheet.merge_range('B19:E19', '', heading7)
    worksheet.merge_range('B20:E20', '', heading7)
    worksheet.merge_range('B21:C21', '', heading7)
    worksheet.merge_range('B22:C22', '', heading7)
    worksheet.merge_range('B23:E23', '', heading7)
    worksheet.merge_range('B24:C24', '', heading7)
    worksheet.merge_range('B25:C25', '', heading7)
    worksheet.merge_range('B26:C26', '', heading7)
    worksheet.merge_range('B27:E27', '', heading7)

    return workbook, worksheet


def define_column_widths_fn(workbook, worksheet):
    """ define and set column widths.
            :param workbook: open workbook object derived from create_workbook_fn function.
            :param worksheet: worksheet object current worksheet derived from create_worksheet_fn.
            :return: workbook: updated workbook object.
            :return worksheet: updated worksheet object."""

    worksheet.set_column('A:A', 45.73)
    worksheet.set_column('B:B', 10.09)
    worksheet.set_column('C:C', 26.91)
    worksheet.set_column('D:D', 26.91)
    worksheet.set_column('E:E', 26.91)

    return workbook, worksheet


def insert_vertical_data_fn(worksheet, row, col, input_list, style, factor):
    """ Insert data (list of lists) vertically.
            :param worksheet: worksheet object current worksheet derived from create_worksheet_fn.
            :param row: integer object containing the row start number for data insertion.
            :param col: integer object containing the column start number for data insertion.
            :param input_list: list of lists object containing data for insertion.
            :param style: style guide variable - heading style passed through function (i.e. (heading1-7).
            :param factor: integer object containing number of rows to skip/not skip."""

    if input_list[0]:
        # Iterate over the data and write it out row by row.
        for item in input_list:
            worksheet.write(row, col, item, style)
            row += factor


def insert_horizontal_data_fn(worksheet, row, col, input_list, style, factor):
    """ Insert data (list of lists) horizontally.
            :param worksheet: worksheet object current worksheet derived from create_worksheet_fn.
            :param row: integer object containing the row start number for data insertion.
            :param col: integer object containing the column start number for data insertion.
            :param input_list: list of lists object containing data for insertion.
            :param style: style guide variable - heading style passed through function (i.e. (heading1-7).
            :param factor: integer object containing number of rows to skip/not skip."""

    if input_list[0]:
        # Iterate over the data and write it out row by row.
        for item in input_list:
            worksheet.write(row, col, item, style)
            col += factor


def main_routine(heading1, heading2, heading3, heading7, workbook, obs_data_list):
    """This scrip creates the site establishment worksheet within th newly created observation workbook.
    :param heading1:
    :param heading2:
    :param heading3:
    :param heading7:
    :param workbook:
    :param obs_data_list:
    :return:
    """
    print('createSiteEstablishmentSheetP.py INITIATED.')

    work_sheet_name = 'Step 1 - Site Establishment'

    # call the create_worksheet_fn function.
    workbook, worksheet = create_worksheet_fn(workbook, work_sheet_name)

    # call the insert_sheet_headings_fn function.
    workbook, worksheet = insert_sheet_headings_fn(workbook, worksheet, heading2, heading3)

    # call the insert_blank_formatted_cells_fn function.
    workbook, worksheet = insert_blank_formatted_cells_fn(workbook, worksheet, heading7)

    # call the merge_cells_fn function.
    workbook, worksheet = merge_cells(workbook, worksheet, work_sheet_name, heading1, heading2, heading7)

    # call the define_column_widths_fn(workbook) function
    workbook, worksheet = define_column_widths_fn(workbook, worksheet)

    establishment_data_list = obs_data_list[0]
    if establishment_data_list[0]:
        print('establishment_data_list: ', establishment_data_list)
        insert_vertical_data_fn(worksheet, 2, 1, establishment_data_list[0], heading7, 1)
    if establishment_data_list[1]:
        # call the insertDataFN function
        insert_vertical_data_fn(worksheet, 8, 1, establishment_data_list[1], heading7, 1)
    if establishment_data_list[2]:
        # call the insertDataFN function
        insert_vertical_data_fn(worksheet, 9, 1, establishment_data_list[2], heading7, 1)
    if establishment_data_list[3]:
        # call the insertDataFN function
        insert_vertical_data_fn(worksheet, 16, 1, establishment_data_list[3], heading7, 1)
    if establishment_data_list[4]:
        # call the insertDataFN function
        insert_vertical_data_fn(worksheet, 20, 4, establishment_data_list[4], heading7, 1)
    if establishment_data_list[5]:
        # call the insertDataFN function
        insert_vertical_data_fn(worksheet, 23, 4, establishment_data_list[5], heading7, 1)

    print('step10_3_site_create_establishment_sheet.py COMPLETE.')
    print('step10_4_site_create_visit_sheet.py initiating..........')

    return insert_vertical_data_fn, insert_horizontal_data_fn

if __name__ == '__main__':
    main_routine()
